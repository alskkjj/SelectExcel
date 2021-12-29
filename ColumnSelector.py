import re, math
import openpyxl

import SelectorConstant
from misc import pyxl_xy, sheet_range


def custom_always_true(s):
    return True


class Selector:
    #    # evaluated value assert
    #    __num_equal_set = set()
    #
    #    # Now number only, expressions and variables in excel aren't support yet
    #    def __num_equals(self, sheet, col):
    #        e_r = re.compile(r'=(?P<expr>.*)')
    #        n_r = re.compile(r'(?P<signature>[\+-]?)(?P<integral>\d+)(?P<dot>\.)?(?P<fraction>(?(dot)\d+))')
    #        filted_rows = set()
    #        for i in range(sheet.max_row):
    #            s = sheet[pyxl_xy(i, col)].value
    #            if s is None:
    #                continue
    #            if s is str and n_r.fullmatch(s):
    #                for r in self.__num_equal_set:
    #                    if type(r) is int and r == int(s):
    #                        filted_rows.add(i)
    #                    elif type(r) is float and float_eq(r, float(s)):
    #                        filted_rows.add(i)
    #            elif s is int:
    #                if any((s == x for x in self.__num_equal_set)):
    #                    filted_rows.add(i)
    #            elif s is float:
    #                if any((s == x for x in self.__num_equal_set)):
    #                    filted_rows.add(i)
    #            else:
    #                print(type(s))
    #                print('\t' + str(s))
    #        return filted_rows

    #    def add_num_equal(self, nm):
    #        if type(nm) is float or type(nm) is int:
    #            self.__num_equal_set.add(nm)
    #        else:
    #            print(type(nm))
    #            raise "Unknow thing add into number equal set"

    # has any match
    __any_match_set = set()

    def __any_match(self, sheet, col):
        filted_rows = set()
        for i in sheet_range(sheet.max_row):
            for r in self.__any_match_set:
                m = r.search(str(sheet[pyxl_xy(i, col)].value))
                if m is not None:
                    filted_rows.add(i)
                    break

        return filted_rows

    def add_any_match(self, mt):
        if type(mt) is str:
            r = re.compile(mt)
            self.__any_match_set.add(r)
        elif type(mt) is re.Pattern:
            self.__any_match_set.add(mt)
        else:
            raise "Unknown thing add into any match set"

    # fullmatch
    __fullmatch_set = set()

    def __fullmatch(self, sheet, col):
        filted_rows = set()
        for i in sheet_range(sheet.max_row):
            for r in self.__fullmatch_set:
                m = r.fullmatch(str(sheet[pyxl_xy(i, col)].value))
                if m is not None:
                    filted_rows.add(i)
                    break
        return filted_rows

    def add_fullmatch(self, mt):
        if type(mt) is str:
            self.__fullmatch_set.add(re.compile(mt))
        elif type(mt) is re.Pattern:
            self.__fullmatch_set.add(mt)
        else:
            raise "Unknown thing add into full match set"

    # custom function: bool(string)
    __funcs_set = set()

    def __custom_function_match(self, sheet, col):
        filted_rows = set()
        for i in sheet_range(sheet.max_row):
            s = str(sheet[pyxl_xy(i, col)].value)
            if any((x(s) for x in self.__funcs_set)):
                filted_rows.add(i)
        return filted_rows

    def add_funcs(self, rep):
        # temporary
        if rep == SelectorConstant.all_selected:
            self.__funcs_set.add(custom_always_true)

    def __init__(self, any_match_set=None, fullmatch_set=None, custom_funcs_set=None):
        self.__any_match_set = any_match_set or set()
        self.__fullmatch_set = fullmatch_set or set()
        self.__funcs_set = custom_funcs_set or set()

    def filter(self, sheet, col):
        filted_rows = set()
        # x+1 means computer's index number (from 0) to users' row number (start from 1)
        filted_rows |= set((x for x in self.__any_match(sheet, col)))
        filted_rows |= set((x for x in self.__fullmatch(sheet, col)))
        filted_rows |= set((x for x in self.__custom_function_match(sheet, col)))
        return filted_rows


class ColumnSelector:
    def __init__(self, column_code, selector: Selector):
        self.column_code = column_code
        self.selector = selector

    def __access_col_code(self):
        return self.column_code

    def single_col_filter(self, sheet):
        return self.selector.filter(sheet, self.__access_col_code())


def test_filters():
    any_match = set()
    any_match.add(re.compile("abc"))
    any_match.add(re.compile("cde"))
    fullmatch = {re.compile("abcde")}
    custom_funcs = set([lambda x: x == "abc"])

    filt = Selector(any_match, fullmatch, custom_funcs)
    filt.add_any_match('强电')
    workbook = openpyxl.load_workbook("assets/强电 - 副本.xlsx")
    sheet = workbook[workbook.sheetnames[0]]
    fted = filt.filter(sheet, 23)
    # print(sheet[pyxl_xy(23, 22)])
    # print(sheet[pyxl_xy(23, chr(ord('A') + 22) )])
    # print(str(sheet[pyxl_xy(22, 22)]))
    # print(sheet.(23, 22))
    assert (fted == {1, 2, 3, 4, 22, 12})
    filt.add_funcs(custom_always_true)
    fted = filt.filter(sheet, 23)
    assert(set(fted) == set(sheet_range(sheet.max_row)))

def test_cao():
    workbook = openpyxl.load_workbook("assets/强电 - 副本07.xlsx")
    sheet = workbook[workbook.sheetnames[0]]
    print(sheet.max_row)
    assert (pyxl_xy(1, 1) == "A1")
    assert (pyxl_xy(5, 26) == "Z5")
