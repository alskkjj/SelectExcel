import openpyxl

from misc import pyxl_xy, sheet_range, tuple_canceler


def product_Decare(s1, s2):
    if s1 is None or len(s1) == 0:
        return s2
    ret = set()
    for x in s1:
        for y in s2:
            ret.add((x, y))
    return ret


# return a dict of element and its row numbers
def group_by_column(sheet, col, selected_row_set=None):
    elements = {'': []}
    if selected_row_set is None:
        selected_row_set = sheet_range(sheet.max_row)
    for i in selected_row_set:
        val = sheet[pyxl_xy(i, col)].value
        elename = ''
        if val is not None:
            elename = str(val)

        if elename not in elements.keys():
            elements[elename] = [i]
        else:
            elements[elename].append(i)
    if len(elements['']) == 0:
        del elements['']
    return elements


def _readable(v):
    if v is None:
        return ''
    else:
        return str(v)


def group_by_columns(sheet, cols, selected_row_set=None):
    _z = ('' for _ in range(len(cols)))
    elements_dict = {_z: []}
    if selected_row_set is None:
        selected_row_set = sheet_range(sheet.max_row)

    for i in selected_row_set:
        ele_k = tuple(_readable(sheet[pyxl_xy(i, col)].value) for col in cols)
        if ele_k not in elements_dict.keys():
            elements_dict[ele_k] = [i]
        else:
            elements_dict[ele_k].append(i)

    if len(elements_dict[_z]) == 0:
        del elements_dict[_z]
    return elements_dict


def test_group_by_column():
    workbook = openpyxl.load_workbook('assets/强电 - 副本07.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    res = group_by_column(sheet, 'D')
    for i in res:
        print(i, res[i])
    res = group_by_column(sheet, 'D', range(26, 30))
    for i in res:
        print(i, res[i])


def test_product_set():
    a = {2, 3, 4}
    b = {'s', 'b', 'c'}
    r = product_Decare(a, b)
    assert (r == {(4, 's'), (2, 'b'), (3, 'b'), (2, 's'), (4, 'c'), (3, 's'), (2, 'c'), (3, 'c'), (4, 'b')})
    a = {(1, 2), (3, 9)}
    b = {'s', 'b', 'c'}
    r = product_Decare(a, b)
    assert (r == {((1, 2), 'b'), ((3, 9), 'c'), ((1, 2), 's'), ((3, 9), 'b'), ((3, 9), 's'), ((1, 2), 'c')})


def test_tuple_canceler():
    a = (1, 2, 3, (3, (444, 5)))
    assert (tuple_canceler(a) == (1, 2, 3, 3, 444, 5))
    a = 1
    assert (tuple_canceler(a) == (1,))


def test_group_by_columns():
    workbook = openpyxl.load_workbook('assets/强电 - 副本07.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    res = group_by_columns(sheet, ['D', 'E'])
    for x in res:
        print(x, res[x])
    res = group_by_columns(sheet, ['I'])
    for x in res:
        print(x, res[x])
