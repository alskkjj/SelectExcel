import re

import openpyxl

from misc import pyxl_xy, float_eq

after_eq_r = re.compile(r'=(?P<expr>.*)')
excel_cord_r = re.compile(r'(?P<col_code>[A-Za-z]+)(?P<row_num>\d+)')


# func(sheet, row, column)
def apply_to_a_column_of_cells(sheet, col, func, rows=None):
    if rows is None:
        rows = range(1, sheet.max_row + 1)

    for i in rows:
        func(sheet, i, col)


def str_quota(s: str):
    return '(' + s + ')'


def treat_pure_number(sheet, row, col):
    clean_rep = '0'
    nm = 0.0
    cell = sheet[pyxl_xy(row, col)]
    if cell.value is None:
        return clean_rep, nm
    try:
        # raw float number, if it's not, then go to exception
        nm = float(cell.value)
        clean_rep = str(cell.value)
    except ValueError:
        return str_quota(str(cell.value)), str_quota(str(nm))

    return clean_rep, nm


def strip_pure_number(sheet, row, col):
    number_r = re.compile(r'/(?P<integral>\d+)(?P<dot>\.)?(?(dot)(?P<fraction>\d+))')
    cell = sheet[pyxl_xy(row, col)]
    nm = 0.0
    clean_rep = ''
    #    if cell.data_type == 'f' or cell.data_type == 'n':
    # pure number
    # we process: pure numbers, zero, null at here, after here they will not be such types
    clean_rep, nm = treat_pure_number(sheet, row, col)
    if type(nm) is float or int:
        return clean_rep, nm

    str_v = str(cell.value)
    mt = after_eq_r.match(str_v)
    # it's not led by =, so we treat it as a normal excel variable expression
    # we process all other strange expressions at here, not led by a =
    # after here they won't be 'led by =' statement
    if mt is None:
        st = str_v
        return '(' + st + ')', '(' + st + ')'

    # check if expr part is excel coordinate format
    # if so do recursive call
    # we intend to process excel cell reference, but it's not comprehensive and not necessary.
    # because the openpyxl has a 'data_only' parameter
    ex_fm = excel_cord_r.fullmatch(mt.group('expr'))
    if ex_fm is not None:
        rn = int(ex_fm.group('row_num'))
        col = ex_fm.group('col_code')
        return strip_pure_number(sheet, rn, col)

    # check if expr part is math expression, which is evaluatable to eval function
    # we intend to process pure the mathematical expressions, but it's not comprehensive and not necessary.
    # because the openpyxl has a 'data_only' parameter
    try:
        nm = eval(mt.group('expr'))
        clean_rep = mt.group('expr')
    except Exception:
        # after eval failed, we consider it's a '=.*' format expression,
        # the contents after = symbol must contain the variable calculation, so we return its expr
        st = mt.group('expr')
        return '(' + st + ')', '(' + st + ')'
    #    elif cell.data_type == 's':
    #    str_v = str(cell.value)

    return clean_rep, nm


def sum_to_a_column_of_cells(sheet, col, rows=None):
    literal_sum = ''
    sum_ = 0.0

    cannot_float_flag = False

    def func(sheet1, row, col1):
        nonlocal literal_sum
        nonlocal sum_
        nonlocal cannot_float_flag
        cell = sheet1[pyxl_xy(row, col1)]
        clean_rep, res = strip_pure_number(sheet1, row, col1)
        if res is None:
            raise ValueError('The column contains non addable element', cell.row, cell.column, clean_rep, res)
        if type(res) is float or type(res) is int:
            if not cannot_float_flag:
                literal_sum += clean_rep + '+'
                sum_ += float(res)
            else:
                literal_sum += clean_rep + '+'
        elif type(res) is str:
            cannot_float_flag = True
            literal_sum += clean_rep + '+'

    apply_to_a_column_of_cells(sheet, col, func, rows)

    if literal_sum.endswith('+'):
        literal_sum = literal_sum[:-1]
    if cannot_float_flag:
        return literal_sum, None 
    else:
        return literal_sum, sum_


def test_apply_to_a_column_of_cells():
    workbook = openpyxl.load_workbook('assets/强电 - 副本07.xlsx')
    sheet = workbook[workbook.sheetnames[0]]
    asstlst = {'=P35', '18', '=1+2', 'h_mezzanine+h_base+h_floor*24'}
    reslst = set()
    apply_to_a_column_of_cells(sheet, 'Q', lambda sh, r, c: reslst.add(str(sheet[pyxl_xy(r, c)].value)),
                               [35, 36, 37, 38])
    assert (reslst == asstlst)

    asstlst = {('((h_mezzanine+h_base+h_floor*24))', '((h_mezzanine+h_base+h_floor*24))'),
               ('18', 18.0),
               ('1+2', 3),
               ('(h_mezzanine+h_base+h_floor*24)', '(h_mezzanine+h_base+h_floor*24)')}
    reslst = set()
    apply_to_a_column_of_cells(sheet, 'Q', lambda sh, r, c: reslst.add((strip_pure_number(sheet, r, c))), [35, 36, 37, 38])


def test_sum_to_a_column_of_cells():
    # sum_to_a_column_of_cells(sheet, col, rows=None):
    workbook = openpyxl.load_workbook('assets/强电 - 副本07.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    # 243.78
    literal_sum, su = sum_to_a_column_of_cells(sheet, 'T', rows=[32, 33, 34, 35])
    assert (float_eq(su, 243.78))
    literal_sum, su = sum_to_a_column_of_cells(sheet, 'H', rows=[5, 6, 7])
    print(literal_sum, su)
    assert(su is None and literal_sum == '(管内穿线)+(管内穿线)+(管内穿线)')
